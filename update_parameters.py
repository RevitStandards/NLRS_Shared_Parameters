#!/usr/bin/env python3
"""
Script om NLRS Shared Parameters bij te werken.
Wijzigingen:
1. NLRS_S_oppervlakte doorsnede: groep 44 → 122
2. NLRS_S_soortelijk gewicht → NLRS_C_soortelijk gewicht (hernoemen)
3. NLRS_S_gewicht per m: groep 105 → 127
"""

import openpyxl
import os

# Configuratie
REPO_DIR = r"C:\Users\marti\Mijn Drive (martijnderiet@mdr-advies.nl)\Github\NLRS_Shared_Parameters"
EXCEL_FILE = os.path.join(REPO_DIR, "NLRS_Shared Parameters - Bronbestand.xlsx")
TOTAAL_FILE = os.path.join(REPO_DIR, "NLRS_Shared Parameters - Totaal.txt")
ALGEMEEN_FILE = os.path.join(REPO_DIR, "NLRS_Shared Parameters - Algemeen.txt")


def update_excel():
    """Update het Excel bestand"""
    print("Updating Excel file...")

    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active

    # Kolom nummers (gebaseerd op headers)
    NAME_COL = 14  # NAME kolom
    GROUP_COL = 17  # GROUP kolom

    changes_made = []

    for row in range(2, ws.max_row + 1):
        param_name = ws.cell(row=row, column=NAME_COL).value

        if param_name == "NLRS_S_oppervlakte doorsnede":
            old_group = ws.cell(row=row, column=GROUP_COL).value
            ws.cell(row=row, column=GROUP_COL).value = 122
            changes_made.append(f"{param_name}: groep {old_group} -> 122")

        elif param_name == "NLRS_S_soortelijk gewicht":
            ws.cell(row=row, column=NAME_COL).value = "NLRS_C_soortelijk gewicht"
            changes_made.append(f"{param_name} -> NLRS_C_soortelijk gewicht")

        elif param_name == "NLRS_S_gewicht per m":
            old_group = ws.cell(row=row, column=GROUP_COL).value
            ws.cell(row=row, column=GROUP_COL).value = 127
            changes_made.append(f"{param_name}: groep {old_group} -> 127")

    wb.save(EXCEL_FILE)
    print(f"Excel updates: {changes_made}")
    return changes_made


def read_utf16_file(filepath):
    """Lees een UTF-16 bestand"""
    with open(filepath, "rb") as f:
        content = f.read()
    return content.decode("utf-16-le")


def write_utf16_file(filepath, content):
    """Schrijf naar een UTF-16 bestand"""
    with open(filepath, "wb") as f:
        f.write(content.encode("utf-16-le"))


def update_text_file(filepath, is_utf16=False):
    """Update een text parameter file"""
    print(f"\nUpdating {os.path.basename(filepath)}...")

    if is_utf16:
        content = read_utf16_file(filepath)
    else:
        with open(filepath, "r", encoding="utf-8") as f:
            content = f.read()

    lines = content.split("\n")
    changes_made = []

    for i, line in enumerate(lines):
        if line.startswith("PARAM"):
            parts = line.split("\t")
            if len(parts) >= 6:
                param_name = parts[2]

                if param_name == "NLRS_S_oppervlakte doorsnede":
                    old_group = parts[5]
                    parts[5] = "122"
                    lines[i] = "\t".join(parts)
                    changes_made.append(f"{param_name}: groep {old_group} -> 122")

                elif param_name == "NLRS_S_soortelijk gewicht":
                    parts[2] = "NLRS_C_soortelijk gewicht"
                    lines[i] = "\t".join(parts)
                    changes_made.append(f"{param_name} -> NLRS_C_soortelijk gewicht")

                elif param_name == "NLRS_S_gewicht per m":
                    old_group = parts[5]
                    parts[5] = "127"
                    lines[i] = "\t".join(parts)
                    changes_made.append(f"{param_name}: groep {old_group} -> 127")

    # Schrijf terug
    new_content = "\n".join(lines)
    if is_utf16:
        write_utf16_file(filepath, new_content)
    else:
        with open(filepath, "w", encoding="utf-8") as f:
            f.write(new_content)

    print(f"Changes in {os.path.basename(filepath)}: {changes_made}")
    return changes_made


def main():
    print("NLRS Shared Parameters Updater")
    print("=" * 50)

    all_changes = []

    # Update Excel
    try:
        changes = update_excel()
        all_changes.extend(changes)
    except Exception as e:
        print(f"Fout bij Excel update: {e}")
        import traceback

        traceback.print_exc()

    # Update Totaal.txt
    try:
        changes = update_text_file(TOTAAL_FILE, is_utf16=False)
        all_changes.extend(changes)
    except Exception as e:
        print(f"Fout bij Totaal.txt update: {e}")
        import traceback

        traceback.print_exc()

    # Update Algemeen.txt (UTF-16)
    try:
        changes = update_text_file(ALGEMEEN_FILE, is_utf16=True)
        all_changes.extend(changes)
    except Exception as e:
        print(f"Fout bij Algemeen.txt update: {e}")
        import traceback

        traceback.print_exc()

    print("\n" + "=" * 50)
    print("Alle wijzigingen:")
    for change in all_changes:
        print(f"  - {change}")


if __name__ == "__main__":
    main()
